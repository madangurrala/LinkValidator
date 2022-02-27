import http.client as httplib
import openpyxl

workbook = openpyxl.load_workbook("links.xlsx")
worksheet = workbook['Sheet1']

resWorkbook = openpyxl.Workbook()
resWorksheet = resWorkbook.active
resWorksheet.title = "Sheet1"
header = [u'URL', u'Response',]
resWorksheet.append(header)
for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=1):
    for cell in row:
        url = cell.value
        connection = httplib.HTTPConnection(url)
        connection.request("HEAD", '')
        response = connection.getresponse().status
        urls = [url, response]
        resWorksheet.append(urls)

resWorkbook.save(filename='responses.xlsx')

# use requests for http/https
